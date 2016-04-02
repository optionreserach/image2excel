from openpyxl import Workbook
from openpyxl.styles import PatternFill,fills
from openpyxl.cell import get_column_letter
from PIL import Image
from pathlib import Path
import os

class image2excel(object):
	
	def __init__(self,image):
		#loading image
		self.absolute_path=Path(image).absolute()
		self.working_dir=self.absolute_path.parent
		os.chdir(str(self.working_dir))
		self.name=self.absolute_path.name.split('.')[0]
		self.img=Image.open(image)
		self.img_width=self.img.size[0]
		self.img_height=self.img.size[1]
		#self.img.thumbnail((self.img_width/(self.img_width//200),300))

	#RGB to hex without alpha
	def _get_color_index(self,tp):
		return '{0:0>2X}{1:0>2X}{2:0>2X}'.format(*tp)

	def to_excel(self):
		#如果宽度大于300像素，对图像进行比例缩小
		if self.img_width>300:
			to_width=self.img_width/(self.img_width//300)
			to_height=self.img_height/(self.img_height//300)
			self.img.thumbnail((to_width,to_height))


		#新建excel文件
		wb=Workbook()
		ws=wb.create_sheet(index=0,title='pixel_picture')


		#使用图片像素填充单元格并调整单元格高度和宽度
		for r in range(self.img.size[1]):
			#该句被移至循环末尾，只有add_cell之后，相应dimension才会增加
			#ws.row_dimensions[r+1].height=2
			for c in range(self.img.size[0]):
				#同理移至末尾
				#ws.column_dimensions[get_column_letter(c+1)].width=0.23
				#获取cell
				_cell=ws.cell(row=r+1,column=c+1)
				#获取hex color index
				_color=self._get_color_index(self.img.getpixel((c,r)))
				_fill=PatternFill(fill_type=fills.FILL_SOLID,start_color=_color)
				_cell.fill=_fill
				ws.column_dimensions[get_column_letter(c+1)].width=0.23
			ws.row_dimensions[r+1].height=2
		#输出excel，与图像同名
		excel_name=self.name+'.xlsx'
		wb.save(excel_name)


def main():
	try:
		print('---在图片路径下生产同名xlsx文件---')
		filename=input('输入图片绝对路径名称：')
		p2e=image2excel(filename)
		p2e.to_excel()
	except Exception as e:
		print(e,"出错啦")

if __name__=='__main__':
	main()

